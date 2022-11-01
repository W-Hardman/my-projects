
import sys
import os
import shutil
from dateutil import parser
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border, colors
from openpyxl.cell import Cell
from openpyxl.styles import NamedStyle
from openpyxl.utils import range_boundaries
import time
import re
from re import sub
from decimal import Decimal
from collections import defaultdict

file = " ".join(sys.argv[1:])

def main():
   
    if len(sys.argv) < 2:
        sys.exit("Usage: python3 automation.py Example Client Data (date).xlsx")
    else:  
        try:
            shutil.copyfile('ClientDeclarationtemplate.xlsx', ('Client Declaration.xlsx'))
        
        except PermissionError:
            sys.exit("Permission Denied")

        except:
            sys.exit("Error occurred while copying file")
    
    copy(file, 'Client Declaration.xlsx')

    clean_client_data()

    clean_declaration_data()

    clean_itemised_billing_internal()

    clean_mta()

    clean_itemised_billing_client()

    clean_file()
    
    rename()



def copy(path1, path2):

    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path1 = file
        path2 = 'Client Declaration.xlsx'

        wb1 = xl.load_workbook(path1, data_only = True)
        ws1 = wb1.worksheets[0]

        wb2 = xl.load_workbook(path2)
        ws2 = wb2.worksheets[0]

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(path2)

    else:
        raise ValueError("No file found")


def clean_client_data():

    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws = wb['Sheet 0']

        for row in ws['O2:O{}'.format(ws.max_row + 1)]:
            for cell in row:
                if cell.value is None or cell.value == "" or any(char.isdigit() for char in str(cell.value)) == False:
                    cell.value = "Unspecified"

        for row in ws['A2:A{}'.format(ws.max_row + 1)]:
            for cell in row:
                orderreference = str(cell.value)
                if not orderreference.startswith('JOTR') and not cell.value is None:
                    redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
                    cell.fill = redFill

        for row in ws['D2:D{}'.format(ws.max_row + 1)]:
            for cell in row:
                email = str(cell.value)
                email_check = re.compile(r"^[^\s@]+@([^\s@.,]+\.)+[^\s@.,]{2,}$")
                if not email_check.match(email) and not cell.value is None:
                    redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
                    cell.fill = redFill

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row = row, column = 12)
            cell.number_format = "dd-mm-yyyy"

        for row in ws['M2:M{}'.format(ws.max_row + 1)]:
            for cell in row:
                if cell.value is None or int(cell.value) % 12 != 0:
                    redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
                    cell.fill = redFill

        for row in ws['P2:Q{}'.format(ws.max_row + 1)]:
            for cell in row:
                if 'Â' in str(cell.value):
                    cell.value = str(cell.value).replace('Â', '')
                if str(cell.value) == "£0.00":
                    cell.value = "£30.00"


        for row in ws['A2:A{}'.format(ws.max_row + 1)]:
            for cell in row:
                if cell.value is None or cell.value == "#REF!":
                    ws.delete_rows(cell.row, 1)


        for row in range(2, ws.max_row + 1):
            if not (ws.cell(row = row, column = 13).value) is None:
                term = int(ws.cell(row = row, column = 13).value)
                if term == 12:
                    yr = "1YR"
                elif term == 24:
                    yr = "2YR"
                elif term == 36:
                    yr = "3YR"
                elif term == 48:
                    yr = "4YR"
                else:
                    yr = "5YR"
                product = str(ws.cell(row = row, column = 14).value)
                prices = str(ws.cell(row = row, column = 17).value)
                values = Decimal(sub(r'[^\d.]', '', prices))

                if "Apple" in product:
                    ws.cell(row = row, column = 18).value = "IR89T" + yr + "APP"

                elif "Apple" not in product:
                    if 0 <= values < 499.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "A"

                    elif 500 < values < 599.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "B"

                    elif 600 < values < 899.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "C"

                    elif 900 < values < 1299.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "D"

                    elif 1300 < values < 1799.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "E"

                    elif 1800 < values < 1999.99:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "F"

                    elif 2000 < values:
                        ws.cell(row = row, column = 18).value = "IR89T" + yr + "G"


        ref_dict = defaultdict(list)
        
        for row in range(2, ws.max_row + 1):
            if not ws.cell(row = row, column = 1).value is None:
                ref = str(ws.cell(row = row, column = 1).value)
                nextref = str(ws.cell(row = row + 1, column = 1).value)

                if ref in ref_dict:
                    ref_dict[ref].append(ws.cell(row = row, column = 18).value)
                elif not ref in ref_dict and ref == nextref:
                    ref_dict[ref] = [ws.cell(row = row, column = 18).value]

        for ref in ref_dict:
            APP = [s for s in ref_dict[ref] if "APP" in s]
            G = [s for s in ref_dict[ref] if "G" in s]
            F = [s for s in ref_dict[ref] if "F" in s]
            E = [s for s in ref_dict[ref] if "E" in s]
            D = [s for s in ref_dict[ref] if "D" in s]
            C = [s for s in ref_dict[ref] if "C" in s]
            B = [s for s in ref_dict[ref] if "B" in s]
            A = [s for s in ref_dict[ref] if "A" in s]
            if len(APP):
                ref_dict[ref] = APP[0]            
            elif len(G):
                ref_dict[ref] = G[0]            
            elif len(F):
                ref_dict[ref] = F[0]            
            elif len(E):
                ref_dict[ref] = E[0]            
            elif len(D):
                ref_dict[ref] = D[0]            
            elif len(C):
                ref_dict[ref] = C[0]            
            elif len(B):
                ref_dict[ref] = B[0]            
            else:
                ref_dict[ref] = A[0]

        for row in range(2, ws.max_row + 1):
            ref = str(ws.cell(row = row, column = 1).value)
            if ref in ref_dict:
                ws.cell(row = row, column = 18).value = ref_dict[ref]


        wb.save(path)

    else:
        raise ValueError("No file found")

def clean_declaration_data():
    
    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws1 = wb['Sheet 2']
        ws2 = wb['Sheet 0']

        for cell in ws2['A:A']:
            ws1.cell(row = cell.row, column = 1, value = cell.value)

        for cell in ws2['B:B']:
            ws1.cell(row = cell.row, column = 3, value = cell.value)

        date = file.rstrip("xlsm")
        date = parser.parse(date, fuzzy=True, dayfirst=True)
        date = str(date)
        date = date[2:-9]
        date = date.split("-")
        date[0], date[-1] = date[-1], date[0]
        date = "".join(date)

        for row in ws1['M2:M{}'.format(ws1.max_row + 1)]:
            for cell in row:
                cell.value = date

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if not str(cell.value).startswith('JOTR'):
                    ws1.delete_rows(cell.row, 1)

        wb.save(path)


def clean_itemised_billing_internal():
    
    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws1 = wb['Sheet 1']
        ws2 = wb['Sheet 2']

        for row in range(2, ws2.max_row + 1):
            for col in range(1, 2):
                ws1.cell(row = row, column = col).value = ws2.cell(row = row, column = col).value

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if not str(cell.value).startswith('JOTR'):
                    ws1.delete_rows(cell.row, 1)

        IR = []
        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:               
                if str(cell.value) in IR:
                    ws1.delete_rows(cell.row, 1)
                else:                   
                    IR.append(cell.value)

        wb.save(path)

def clean_mta():

    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws1 = wb['Sheet 3']
        ws2 = wb['Sheet 2']

        for row in range(2, ws2.max_row + 1):
            for col in range(1, 2):
                ws1.cell(row = row, column = col).value = ws2.cell(row = row, column = col).value

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if not str(cell.value).startswith('JOTR'):
                    ws1.delete_rows(cell.row, 1)

        wb.save(path)

def clean_itemised_billing_client():
    
    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1)

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws1 = wb['Sheet 4']
        ws2 = wb['Sheet 2']

        for row in range(2, ws2.max_row + 1):
            for col in range(1, 2):
                ws1.cell(row = row, column = col).value = ws2.cell(row = row, column = col).value

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if not str(cell.value).startswith('JOTR'):
                    ws1.delete_rows(cell.row, 1)

        IR = []
        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if str(cell.value) in IR:
                    ws1.delete_rows(cell.row, 1)
                else:                   
                    IR.append(cell.value)

        wb.save(path)

def clean_file():

    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1) 

    if os.path.isfile('Client Declaration.xlsx') == True:
        path = 'Client Declaration.xlsx'
        wb = xl.load_workbook(path)
        ws1 = wb['Sheet 5']
        ws2 = wb['Sheet 2']

        for row in range(2, ws2.max_row + 1):
            for col in range(1, 2):
                ws1.cell(row = row, column = col).value = ws2.cell(row = row, column = col).value

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                cell.value = str(cell.value).strip('JOTR')

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if str(cell.value) == "0" or cell.value is None or cell.value == "" or any(char.isdigit() for char in str(cell.value)) == False:
                    ws1.delete_rows(cell.row, 1)

        for row in ws1['A2:A{}'.format(ws1.max_row + 1)]:
            for cell in row:
                if not cell.value is None:
                    cell.value = int(cell.value)

        wb.save(path)

def rename():
    
    date = file.rstrip("xlsm")
    date = parser.parse(date, fuzzy=True, dayfirst=True)
    date = str(date)
    date = date[2:-9]
    date = date.split("-")
    date[0], date[-1] = date[-1], date[0]
    date = "".join(date)
    
    while os.path.isfile('Client Declaration.xlsx') == False:
        time.sleep(1)

    if os.path.isfile('Client Declaration.xlsx') == True:
        newfile = 'Client Declaration' + f' {date}' + '.xlsx'
        os.rename('Client Declaration.xlsx', newfile)

    else:
        raise ValueError("No file found")

if __name__ == "__main__":
    main()
